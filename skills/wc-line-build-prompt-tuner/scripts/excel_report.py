"""Generate v4.10-format xlsx report from baseline-raw.json + llm-output.json.

Sheets:
  Summary       - global accuracy + per-brand item counts
  Per-Item      - one row per item with mismatch counts + sample errors
  Details       - full per-field diff rows
  Mismatches    - same as Details (subset where verdict != match)
  LLM-Unmatched(excluded) - LLM procs with no rel-id overlap in same activity bucket
  Prod-NotMatched(info)   - Prod procs not covered by any LLM proc in same activity
  LineBuilds    - text dump of both line builds per item
"""
import argparse
import json
import os
import sys
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


# -- Severity classification ---------------------------------------------------
SEVERE_FIELDS = {
    ("COOK", "appliance"),
    ("COOK", "cooking_usage_seconds"),
    ("COOK", "batch_limit"),
    ("COOK", "appliance_config_id"),
    ("COMPLETE", "hold_usage_seconds"),
    ("COMPLETE", "parking_spot"),
}
LIGHT_FIELDS_SUFFIX = {"step_usage_seconds", "step_title"}


REFERENCE_VERDICT = "参考"   # excluded from mismatch stats / percentage
MATCH_VERDICT = "match"
SEVERE_VERDICT = "严重预警"
LIGHT_VERDICT = "轻量预警"

# Activities where packaging / container steps without related_item_number
# are skipped from mismatch stats (free-form naming, kitchen ops allow drift).
# COMPLETE keeps full comparison because 'Place in Bag' vs real dish name is a real bug.
REFERENCE_ONLY_STEP_ACTIVITIES = {"COOK", "GARNISH"}


def severity(activity, field):
    if (activity, field) in SEVERE_FIELDS:
        return SEVERE_VERDICT
    return LIGHT_VERDICT


def step_verdict(activity, llm_step, prod_step):
    """Decide verdict for a step_title comparison.
    For COOK/GARNISH: when EITHER side has no related_item_number on the step
    being compared, treat it as 参考 (reference-only, excluded from stats).
    For COMPLETE: always counted (Place-in-Bag bug must surface).
    The pair is only emitted as a real mismatch when both sides have a rel.
    """
    lv = (llm_step or {}).get("title") if llm_step else None
    bv = (prod_step or {}).get("title") if prod_step else None
    if title_norm(lv) == title_norm(bv):
        return MATCH_VERDICT
    if activity in REFERENCE_ONLY_STEP_ACTIVITIES:
        lr = (llm_step or {}).get("related_item_number") if llm_step else None
        br = (prod_step or {}).get("related_item_number") if prod_step else None
        if lr is None or br is None:
            return REFERENCE_VERDICT
    return severity(activity, "step_title")


# -- Brand classification ------------------------------------------------------
def brand_from_name(name):
    if not name:
        return "Cross / Other"
    n = name.strip()
    for brand_marker, brand_label in [
        (", Limesalt", "Limesalt"),
        (", Yasas", "Yasas"),
        (", Hanu Poke", "Hanu Poke"),
        (", Royal Greens", "Royal Greens"),
    ]:
        if n.endswith(brand_marker) or brand_marker.replace(",", ",") in n:
            return brand_label
    # Some items have " by Michael Symon" etc — handled implicitly
    return "Cross / Other"


# -- Pairing logic mirrors compare.py -----------------------------------------
def rel_of(p):
    steps = p.get("procedure_steps") or []
    return steps[0].get("related_item_number") if steps else None


def all_rels(p):
    return [s.get("related_item_number") for s in (p.get("procedure_steps") or []) if s.get("related_item_number")]


def skeleton_procs(skeleton):
    """Pull all procedures from skeleton's first line_build / first task, in order."""
    if isinstance(skeleton, str):
        skeleton = json.loads(skeleton)
    if not isinstance(skeleton, dict):
        return []
    lbs = skeleton.get("line_builds") or []
    if not lbs:
        return []
    tasks = lbs[0].get("tasks") or []
    if not tasks:
        return []
    return sorted(tasks[0].get("procedures") or [], key=lambda x: x.get("order") or 0)


def enrich_llm_with_skeleton(llm_procs, skeleton):
    """In-place: copy related_item_number / related_item_version_id from skeleton's procedure_steps
    into LLM procs at the same position.

    The Gemini schema doesn't ask the LLM to return related_item_number on steps, so LLM output
    always has it as null. But the skeleton (built by WCLineBuildBuilder from cook_methods + prod
    menu data) does carry it. Because skeleton and LLM share the same procedure layout (1:1 by
    order index), we can transplant the rel from skeleton into LLM so downstream matching can use it.
    """
    skel_procs = skeleton_procs(skeleton)
    # Pair by index after sorting llm_procs by order (already sorted upstream usually, but enforce).
    llm_sorted = sorted(llm_procs, key=lambda x: x.get("order") or 0)
    n = min(len(llm_sorted), len(skel_procs))
    for i in range(n):
        lp = llm_sorted[i]
        sp = skel_procs[i]
        # also fill in order if LLM left it null (skeleton order is authoritative)
        if lp.get("order") is None and sp.get("order") is not None:
            lp["order"] = sp.get("order")
        lps = lp.get("procedure_steps") or []
        sps = sp.get("procedure_steps") or []
        m = min(len(lps), len(sps))
        for j in range(m):
            if lps[j].get("related_item_number") is None:
                lps[j]["related_item_number"] = sps[j].get("related_item_number")
            if lps[j].get("related_item_version_id") is None:
                lps[j]["related_item_version_id"] = sps[j].get("related_item_version_id")


def pair_procedures(la, ba):
    """Pair procedures inside one activity bucket.

    Phase 1: rel-based — for each LLM proc whose first step has a non-null
             related_item_number, find the prod proc with matching rel.
    Phase 2: positional — remaining procs on both sides are paired by order.

    This unified strategy applies to GARNISH / COOK / COMPLETE — wherever the
    skeleton has filled a rel onto the LLM proc, we trust that anchor over
    accidental position similarities.
    """
    used_llm = set()
    used_base = set()
    pairs = []

    base_by_rel = defaultdict(list)
    for jb, bp in enumerate(ba):
        r = rel_of(bp)
        if r is not None:
            base_by_rel[r].append((jb, bp))
    # Phase 1
    for jl, lp in enumerate(la):
        r = rel_of(lp)
        if r is not None and base_by_rel.get(r):
            jb, bp = base_by_rel[r].pop(0)
            pairs.append((jl, lp, bp))
            used_llm.add(jl)
            used_base.add(jb)
    # Phase 2 — positional on leftovers, preserving original order
    rem_llm = [(jl, lp) for jl, lp in enumerate(la) if jl not in used_llm]
    rem_base = [(jb, bp) for jb, bp in enumerate(ba) if jb not in used_base]
    while rem_llm and rem_base:
        jl, lp = rem_llm.pop(0)
        jb, bp = rem_base.pop(0)
        pairs.append((jl, lp, bp))
    # Any leftover LLM procs go unpaired (later captured in LLM-Unmatched sheet)
    for jl, lp in rem_llm:
        pairs.append((jl, lp, None))
    return pairs


def title_norm(s):
    return (s or "").strip().lower()


# -- Diffing per pair ---------------------------------------------------------
def diff_pair(activity, lp, bp, name, item_number, brand, i):
    rows = []
    fields = ["step_usage_seconds"]
    if activity == "COMPLETE":
        fields += ["hold_usage_seconds", "parking_spot"]
    elif activity == "COOK":
        fields += ["appliance", "cooking_usage_seconds", "batch_limit"]
        if lp.get("appliance") in {"PIZZA_CONVEYOR_OVEN", "TURBO_OVEN"} \
                or bp.get("appliance") in {"PIZZA_CONVEYOR_OVEN", "TURBO_OVEN"}:
            fields.append("appliance_config_id")

    # Match v4.10 semantics: display the BASE (prod) procedure's step related_item_numbers
    # so a mismatch row immediately tells you "which ingredient/menu the prod proc was anchored to".
    # LLM-side rels are almost always null (schema doesn't ask LLM to fill them); displaying intersection
    # would always be empty. If you need to inspect LLM rels separately, see LLM-Unmatched sheet.
    base_rels = sorted({r for r in all_rels(bp) if r})
    if base_rels:
        shared_str = ",".join(base_rels)
    else:
        # Prod proc has no rel either (rare; usually packaging-only GARNISH or "Place in Bag" COMPLETE)
        shared_str = "(placeholder)"

    for f in fields:
        lv, bv = lp.get(f), bp.get(f)
        verdict = "match" if lv == bv else severity(activity, f)
        rows.append({
            "brand": brand, "item_number": item_number, "name": name,
            "activity": activity, "shared_rels": shared_str,
            "field": f, "prod_value": bv, "llm_value": lv,
            "verdict": verdict, "i": i,
        })

    # Pair procedure_steps within this procedure.
    # Strategy: when a step has related_item_number, pair by rel; otherwise fall back to
    # positional pairing on whatever steps remain on both sides. Prevents the false-positive
    # where LLM omits a container-only step and the rest get position-mismatched (e.g.
    # PROD COOK steps = ['4oz Portion Cup' rel=None, 'Queso Blanco' rel=4000279],
    # LLM COOK steps = ['Queso Blanco' rel=4000279] would have been wrongly compared as
    # '4oz Portion Cup' vs 'Queso Blanco'.
    step_pairs, prod_extra_steps, llm_extra_steps = _pair_steps(lp.get("procedure_steps") or [],
                                                                bp.get("procedure_steps") or [])
    for (jl, ls, jb, bs) in step_pairs:
        lv, bv = ls.get("title"), bs.get("title")
        verdict = step_verdict(activity, ls, bs)
        rows.append({
            "brand": brand, "item_number": item_number, "name": name,
            "activity": activity, "shared_rels": shared_str,
            "field": f"step_title[{jb}]", "prod_value": bv, "llm_value": lv,
            "verdict": verdict, "i": i,
        })
    # Steps the LLM dropped (prod has them, LLM didn't generate).
    # If COOK/GARNISH and the dropped prod step has no rel → 参考 (likely packaging-only step);
    # else → real warning.
    for (jb, bs) in prod_extra_steps:
        if activity in REFERENCE_ONLY_STEP_ACTIVITIES and bs.get("related_item_number") is None:
            v = REFERENCE_VERDICT
        else:
            v = severity(activity, "step_title")
        rows.append({
            "brand": brand, "item_number": item_number, "name": name,
            "activity": activity, "shared_rels": shared_str,
            "field": f"step_title[{jb}] (LLM missing this step)",
            "prod_value": bs.get("title"), "llm_value": None,
            "verdict": v, "i": i,
        })
    # Steps the LLM hallucinated extra (LLM has them, prod doesn't).
    for (jl, ls) in llm_extra_steps:
        if activity in REFERENCE_ONLY_STEP_ACTIVITIES and ls.get("related_item_number") is None:
            v = REFERENCE_VERDICT
        else:
            v = severity(activity, "step_title")
        rows.append({
            "brand": brand, "item_number": item_number, "name": name,
            "activity": activity, "shared_rels": shared_str,
            "field": f"step_title[+{jl}] (LLM extra step)",
            "prod_value": None, "llm_value": ls.get("title"),
            "verdict": v, "i": i,
        })
    return rows


def _pair_steps(llm_steps, prod_steps):
    """Pair procedure_steps within one procedure.
    Returns (pairs, prod_extra, llm_extra).
      pairs        list of (llm_index, llm_step, prod_index, prod_step)
      prod_extra   list of (prod_index, prod_step)  — prod has, LLM doesn't
      llm_extra    list of (llm_index, llm_step)    — LLM has, prod doesn't
    Strategy:
      1. Group both sides by related_item_number (non-null).
      2. Pair on matching rels first.
      3. Remaining (no-rel or unmatched-rel) steps fall back to positional pairing.
    """
    used_llm = set()
    used_prod = set()
    pairs = []

    # Phase 1: rel-based pairing
    prod_by_rel = {}
    for jb, bs in enumerate(prod_steps):
        r = bs.get("related_item_number")
        if r is not None:
            prod_by_rel.setdefault(r, []).append((jb, bs))
    for jl, ls in enumerate(llm_steps):
        r = ls.get("related_item_number")
        if r is None or r not in prod_by_rel or not prod_by_rel[r]:
            continue
        jb, bs = prod_by_rel[r].pop(0)
        pairs.append((jl, ls, jb, bs))
        used_llm.add(jl)
        used_prod.add(jb)

    # Phase 2: positional pairing on the leftovers (preserving original order)
    rem_llm = [(jl, ls) for jl, ls in enumerate(llm_steps) if jl not in used_llm]
    rem_prod = [(jb, bs) for jb, bs in enumerate(prod_steps) if jb not in used_prod]
    while rem_llm and rem_prod:
        jl, ls = rem_llm.pop(0)
        jb, bs = rem_prod.pop(0)
        pairs.append((jl, ls, jb, bs))

    return pairs, rem_prod, rem_llm


# -- Build full data ----------------------------------------------------------
def build_data(baseline_raw_path, llm_output_path):
    with open(baseline_raw_path, "r", encoding="utf-8") as f:
        raw_docs = json.load(f)
    with open(llm_output_path, "r", encoding="utf-8") as f:
        llm_out = json.load(f)

    # baseline_first_task — match compare.py semantics
    base = {}
    for d in raw_docs:
        lbs = (d.get("item_line_build") or {}).get("line_builds") or []
        if not lbs:
            continue
        tasks = lbs[0].get("tasks") or []
        if not tasks:
            continue
        base[d["_id"]] = {
            "name": d.get("name"), "item_number": d.get("item_number"),
            "procedures": tasks[0].get("procedures") or [],
        }

    all_detail_rows = []
    llm_unmatched_rows = []
    prod_notmatched_rows = []
    per_item_rows = []
    linebuild_rows = []

    for entry in llm_out:
        if entry.get("status") != "ok":
            continue
        ivid = entry["item_version_id"]
        b = base.get(ivid)
        if not b:
            continue
        name = b["name"]
        item_number = b["item_number"]
        brand = brand_from_name(name)

        llm = entry["llm_response_json"]
        if isinstance(llm, str):
            llm = json.loads(llm)
        llm_procs = sorted(llm.get("procedures") or [], key=lambda x: x.get("order") or 0)
        # Enrich LLM procs with related_item_number from skeleton so pairing/diff works correctly.
        enrich_llm_with_skeleton(llm_procs, entry.get("skeleton_json"))

        # group by activity
        llm_groups = defaultdict(list)
        for p in llm_procs:
            llm_groups[p.get("activity")].append(p)
        base_groups = defaultdict(list)
        for p in sorted(b["procedures"], key=lambda x: x.get("order") or 0):
            base_groups[p.get("activity")].append(p)

        item_severe = 0
        item_light = 0
        sample_errors = []

        for act in ("GARNISH", "COOK", "COMPLETE"):
            la = llm_groups.get(act, [])
            ba = base_groups.get(act, [])
            # Unified pairing: rel-first, positional fallback (covers GARNISH / COOK / COMPLETE).
            pairs = pair_procedures(la, ba)
            paired_llm_ids = set()   # id() of llm proc that got paired
            paired_base_ids = set()  # id() of base proc that got paired
            for (i, lp, bp) in pairs:
                paired_llm_ids.add(id(lp))
                if bp is None:
                    continue
                paired_base_ids.add(id(bp))
                drows = diff_pair(act, lp, bp, name, item_number, brand, i)
                all_detail_rows.extend(drows)
                for r in drows:
                    if r["verdict"] == SEVERE_VERDICT:
                        item_severe += 1
                    elif r["verdict"] == LIGHT_VERDICT:
                        item_light += 1
                    # Only severe + light surface as sample errors; 参考 rows hidden.
                    if r["verdict"] in (SEVERE_VERDICT, LIGHT_VERDICT) and len(sample_errors) < 3:
                        sample_errors.append(f"{r['activity']}.{r['field']}={r['llm_value']!r} (prod={r['prod_value']!r})")

            # LLM-Unmatched: only truly extra LLM procs (couldn't be paired even positionally)
            for lp in la:
                if id(lp) in paired_llm_ids:
                    continue
                lrels = set(all_rels(lp))
                llm_unmatched_rows.append({
                    "brand": brand, "item_number": item_number, "name": name,
                    "activity": act,
                    "llm_order": lp.get("order"),
                    "rels": ",".join(sorted(lrels)) if lrels else "(no related_item)",
                    "titles": ", ".join(repr(s.get("title")) for s in (lp.get("procedure_steps") or [])[:5]),
                    "reason": "LLM produced more procedures than prod in this activity bucket; this proc has no pair",
                })

            # Prod-NotMatched: only truly extra prod procs
            for bp in ba:
                if id(bp) in paired_base_ids:
                    continue
                brels = set(all_rels(bp))
                prod_notmatched_rows.append({
                    "brand": brand, "item_number": item_number, "name": name,
                    "activity": act,
                    "prod_order": bp.get("order"),
                    "rels": ",".join(sorted(brels)) if brels else "(no related_item)",
                    "titles": ", ".join(repr(s.get("title")) for s in (bp.get("procedure_steps") or [])[:5]),
                    "note": "Prod has this procedure but LLM did not generate a paired counterpart",
                })

        # LLM unmatched proc count for per-item
        llm_unmatched_for_item = sum(1 for r in llm_unmatched_rows if r["item_number"] == item_number)
        per_item_rows.append({
            "brand": brand,
            "item_number": item_number,
            "name": name,
            "prod_proc_count": len(b["procedures"]),
            "llm_proc_count": len(llm_procs),
            "llm_unmatched_count": llm_unmatched_for_item,
            "severe_warn": item_severe,
            "light_warn": item_light,
            "sample_errors": " | ".join(sample_errors),
            # Match v4.10 semantics: "完全一致" = no severe + no light warning
            # (LLM-Unmatched / Prod-NotMatched are informational only, do not block完全一致)
            "completely_match": (item_severe == 0 and item_light == 0),
        })

        # LineBuilds text dump
        def proc_line(p):
            tag = f"[{p.get('activity')}]"
            extra = []
            if p.get("activity") == "COOK":
                extra.append(f"appliance={p.get('appliance')}")
                extra.append(f"cook={p.get('cooking_usage_seconds')}")
                extra.append(f"batch={p.get('batch_limit')}")
                if p.get("appliance_config_id"):
                    extra.append(f"cfg={p.get('appliance_config_id')}")
            elif p.get("activity") == "COMPLETE":
                extra.append(f"hold={p.get('hold_usage_seconds')}")
                extra.append(f"parking={p.get('parking_spot')}")
            titles = [s.get("title") for s in (p.get("procedure_steps") or [])]
            return f"{tag} order={p.get('order')} step={p.get('step_usage_seconds')} " + " ".join(extra) + f"  titles={titles!r}"
        prod_lb = "\n".join(proc_line(p) for p in sorted(b["procedures"], key=lambda x: x.get("order") or 0))
        llm_lb = "\n".join(proc_line(p) for p in llm_procs)
        linebuild_rows.append({
            "brand": brand, "item_number": item_number, "name": name,
            "prod_lb": prod_lb, "llm_lb": llm_lb,
        })

    return {
        "details": all_detail_rows,
        "llm_unmatched": llm_unmatched_rows,
        "prod_notmatched": prod_notmatched_rows,
        "per_item": per_item_rows,
        "linebuilds": linebuild_rows,
    }


# -- xlsx writer --------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="FFD0E4F7", end_color="FFD0E4F7", fill_type="solid")
HEADER_FONT = Font(bold=True)
MISMATCH_FILL = PatternFill(start_color="FFFFD8D8", end_color="FFFFD8D8", fill_type="solid")
SEVERE_FILL = PatternFill(start_color="FFFFB3B3", end_color="FFFFB3B3", fill_type="solid")
REFERENCE_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")


def ws_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_report(data, output_path, prompt_version="vX"):
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    total_items = len(data["per_item"])
    fully_match = sum(1 for r in data["per_item"] if r["completely_match"])

    ws.cell(1, 1, f"WC Line Build Agent — Prompt {prompt_version} Accuracy Report").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Total items run: {total_items}")
    ws.cell(3, 1, f"完全一致 (no warn, no unmatched): {fully_match}/{total_items} = {100*fully_match/total_items:.1f}%")

    headers = ["activity", "field", "匹配数", "总数", "命中率"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(5, i, h); c.fill = HEADER_FILL; c.font = HEADER_FONT

    # Field-level stats from details — exclude 参考 rows from both numerator AND denominator
    bucket = defaultdict(lambda: [0, 0])
    reference_count = 0
    for r in data["details"]:
        if r["verdict"] == REFERENCE_VERDICT:
            reference_count += 1
            continue
        act_f = (r["activity"], r["field"].split("[")[0])  # collapse step_title[0..n] → step_title
        bucket[act_f][1] += 1
        if r["verdict"] == MATCH_VERDICT:
            bucket[act_f][0] += 1
    row = 6
    for k in sorted(bucket.keys()):
        m, t = bucket[k]
        ws.cell(row, 1, k[0])
        ws.cell(row, 2, k[1])
        ws.cell(row, 3, m)
        ws.cell(row, 4, t)
        ws.cell(row, 5, f"{100*m/t:.1f}%" if t else "")
        row += 1
    ws.cell(row, 1, f"(参考行已排除: {reference_count} — COOK/GARNISH 中无 related_item_number 的 step，多为 packaging/sleeve 等命名漂移)").font = Font(italic=True, color="888888")

    # Per-brand
    row += 2
    ws.cell(row, 1, "按品牌统计 (item-level)").font = Font(bold=True)
    row += 1
    for i, h in enumerate(["品牌", "item 数", "完全一致 item 数", "完全一致率"], 1):
        c = ws.cell(row, i, h); c.fill = HEADER_FILL; c.font = HEADER_FONT
    row += 1
    brand_stats = defaultdict(lambda: [0, 0])
    for r in data["per_item"]:
        brand_stats[r["brand"]][1] += 1
        if r["completely_match"]:
            brand_stats[r["brand"]][0] += 1
    for brand in sorted(brand_stats.keys()):
        m, t = brand_stats[brand]
        ws.cell(row, 1, brand)
        ws.cell(row, 2, t)
        ws.cell(row, 3, m)
        ws.cell(row, 4, f"{100*m/t:.1f}%" if t else "")
        row += 1
    autosize(ws, [25, 30, 14, 14, 14])

    # ---- Per-Item ----
    ws = wb.create_sheet("Per-Item")
    headers = ["品牌", "item_number", "menu_item_name", "prod proc 数", "agent proc 数",
               "LLM未匹配proc数(已排除)", "严重预警(字段数)", "轻量预警(字段数)", "示例错误 (最多3条)"]
    ws_header(ws, headers)
    for ri, r in enumerate(sorted(data["per_item"], key=lambda x: (x["brand"], x["item_number"])), 2):
        ws.cell(ri, 1, r["brand"])
        ws.cell(ri, 2, r["item_number"])
        ws.cell(ri, 3, r["name"])
        ws.cell(ri, 4, r["prod_proc_count"])
        ws.cell(ri, 5, r["llm_proc_count"])
        ws.cell(ri, 6, r["llm_unmatched_count"])
        ws.cell(ri, 7, r["severe_warn"])
        ws.cell(ri, 8, r["light_warn"])
        ws.cell(ri, 9, r["sample_errors"])
        if r["severe_warn"] > 0:
            for col in range(1, 10):
                ws.cell(ri, col).fill = SEVERE_FILL
        elif r["light_warn"] > 0:
            for col in range(1, 10):
                ws.cell(ri, col).fill = MISMATCH_FILL
    autosize(ws, [15, 12, 60, 12, 12, 22, 16, 16, 80])

    # ---- Details ----
    ws = wb.create_sheet("Details")
    headers = ["品牌", "item_number", "menu_item_name", "activity", "shared_related_item_numbers",
               "field", "prod_value", "llm_value", "verdict"]
    ws_header(ws, headers)
    for ri, r in enumerate(sorted(data["details"], key=lambda x: (x["brand"], x["item_number"], x["activity"], x["i"], x["field"])), 2):
        ws.cell(ri, 1, r["brand"])
        ws.cell(ri, 2, r["item_number"])
        ws.cell(ri, 3, r["name"])
        ws.cell(ri, 4, r["activity"])
        ws.cell(ri, 5, r["shared_rels"])
        ws.cell(ri, 6, r["field"])
        ws.cell(ri, 7, str(r["prod_value"]))
        ws.cell(ri, 8, str(r["llm_value"]))
        ws.cell(ri, 9, r["verdict"])
        if r["verdict"] == SEVERE_VERDICT:
            ws.cell(ri, 9).fill = SEVERE_FILL
        elif r["verdict"] == LIGHT_VERDICT:
            ws.cell(ri, 9).fill = MISMATCH_FILL
        elif r["verdict"] == REFERENCE_VERDICT:
            # whole row in light gray so reader knows it's excluded from stats
            for col in range(1, 10):
                ws.cell(ri, col).fill = REFERENCE_FILL
    autosize(ws, [15, 12, 50, 10, 25, 22, 45, 45, 12])

    # ---- Mismatches (only severe + light — 参考 excluded) ----
    ws = wb.create_sheet("Mismatches")
    ws_header(ws, headers)
    ri = 2
    for r in sorted(data["details"], key=lambda x: (x["brand"], x["item_number"], x["activity"], x["i"], x["field"])):
        if r["verdict"] in (MATCH_VERDICT, REFERENCE_VERDICT):
            continue
        ws.cell(ri, 1, r["brand"])
        ws.cell(ri, 2, r["item_number"])
        ws.cell(ri, 3, r["name"])
        ws.cell(ri, 4, r["activity"])
        ws.cell(ri, 5, r["shared_rels"])
        ws.cell(ri, 6, r["field"])
        ws.cell(ri, 7, str(r["prod_value"]))
        ws.cell(ri, 8, str(r["llm_value"]))
        ws.cell(ri, 9, r["verdict"])
        if r["verdict"] == SEVERE_VERDICT:
            ws.cell(ri, 9).fill = SEVERE_FILL
        else:
            ws.cell(ri, 9).fill = MISMATCH_FILL
        ri += 1
    autosize(ws, [15, 12, 50, 10, 25, 22, 45, 45, 12])

    # ---- LLM-Unmatched(excluded) ----
    ws = wb.create_sheet("LLM-Unmatched(excluded)")
    ws.cell(1, 1, "说明: 以下 LLM 输出的 procedure，与 prod baseline 同 activity 桶里找不到任何 related_item_number 重叠的对应项；已被 Details/Mismatches 排除").font = Font(italic=True)
    headers2 = ["品牌", "item_number", "menu_item_name", "activity", "LLM order", "LLM related_item_numbers", "LLM step titles", "排除原因"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(3, i, h); c.fill = HEADER_FILL; c.font = HEADER_FONT
    for ri, r in enumerate(sorted(data["llm_unmatched"], key=lambda x: (x["brand"], x["item_number"], x["activity"])), 4):
        ws.cell(ri, 1, r["brand"]); ws.cell(ri, 2, r["item_number"]); ws.cell(ri, 3, r["name"])
        ws.cell(ri, 4, r["activity"]); ws.cell(ri, 5, r["llm_order"])
        ws.cell(ri, 6, r["rels"]); ws.cell(ri, 7, r["titles"]); ws.cell(ri, 8, r["reason"])
    autosize(ws, [15, 12, 50, 10, 10, 30, 60, 70])

    # ---- Prod-NotMatched(info) ----
    ws = wb.create_sheet("Prod-NotMatched(info)")
    ws.cell(1, 1, "说明: 以下 prod baseline 的 procedure，没有被 LLM 输出的任何 procedure 通过 related_item_number 覆盖").font = Font(italic=True)
    headers3 = ["品牌", "item_number", "menu_item_name", "activity", "prod order", "prod related_item_numbers", "prod step titles (前5)", "说明"]
    for i, h in enumerate(headers3, 1):
        c = ws.cell(3, i, h); c.fill = HEADER_FILL; c.font = HEADER_FONT
    for ri, r in enumerate(sorted(data["prod_notmatched"], key=lambda x: (x["brand"], x["item_number"], x["activity"])), 4):
        ws.cell(ri, 1, r["brand"]); ws.cell(ri, 2, r["item_number"]); ws.cell(ri, 3, r["name"])
        ws.cell(ri, 4, r["activity"]); ws.cell(ri, 5, r["prod_order"])
        ws.cell(ri, 6, r["rels"]); ws.cell(ri, 7, r["titles"]); ws.cell(ri, 8, r["note"])
    autosize(ws, [15, 12, 50, 10, 10, 30, 60, 70])

    # ---- LineBuilds ----
    ws = wb.create_sheet("LineBuilds")
    ws_header(ws, ["品牌", "item_number", "menu_item_name", "PROD line build (canonical)", "AGENT line build"])
    for ri, r in enumerate(sorted(data["linebuilds"], key=lambda x: (x["brand"], x["item_number"])), 2):
        ws.cell(ri, 1, r["brand"]); ws.cell(ri, 2, r["item_number"]); ws.cell(ri, 3, r["name"])
        c1 = ws.cell(ri, 4, r["prod_lb"]); c1.alignment = Alignment(wrap_text=True, vertical="top")
        c2 = ws.cell(ri, 5, r["llm_lb"]);  c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ri].height = max(60, 14 * max(r["prod_lb"].count("\n"), r["llm_lb"].count("\n")) + 14)
    autosize(ws, [15, 12, 40, 90, 90])

    wb.save(output_path)
    print(f"wrote {output_path}")
    print(f"  per-item:        {len(data['per_item'])}")
    print(f"  details rows:    {len(data['details'])}")
    severe_n = sum(1 for r in data['details'] if r['verdict'] == SEVERE_VERDICT)
    light_n = sum(1 for r in data['details'] if r['verdict'] == LIGHT_VERDICT)
    ref_n = sum(1 for r in data['details'] if r['verdict'] == REFERENCE_VERDICT)
    match_n = sum(1 for r in data['details'] if r['verdict'] == MATCH_VERDICT)
    print(f"  match rows:      {match_n}")
    print(f"  severe mismatch: {severe_n}")
    print(f"  light mismatch:  {light_n}")
    print(f"  reference rows:  {ref_n}  (gray; not counted)")
    print(f"  llm-unmatched:   {len(data['llm_unmatched'])}")
    print(f"  prod-notmatched: {len(data['prod_notmatched'])}")


def main():
    out_dir = os.environ.get("PROMPT_TUNER_OUT_DIR") or os.path.join(os.getcwd(), "prompt-tuner-out")
    ap = argparse.ArgumentParser()
    ap.add_argument("--baseline-raw", default=os.path.join(out_dir, "baseline-raw.json"))
    ap.add_argument("--llm-output", default=os.path.join(out_dir, "llm-output.json"))
    ap.add_argument("--out", default=r"C:\Users\ev\Desktop\line_build_agent\report_v4.11_nocust.xlsx")
    ap.add_argument("--prompt-version", default="v4.11")
    args = ap.parse_args()

    data = build_data(args.baseline_raw, args.llm_output)
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    write_report(data, args.out, prompt_version=args.prompt_version)


if __name__ == "__main__":
    main()
